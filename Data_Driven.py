from selenium import webdriver
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By

# url parameters
url = "https://www.saucedemo.com/"

# Load the excel sheet
workbook = load_workbook('Test_data.xlsx')

# Select the sheet
sheet = workbook.active

# Open and maximiza browser
driver = webdriver.Chrome()
driver.maximize_window()

# Condition to attempt login with each username/password
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]
    # Access url
    driver.get(url)
    time.sleep(3)

    # Locate and type in username box element
    driver.find_element(By.ID,"user-name").send_keys(username)

    # Locate and type in password box element
    driver.find_element(By.ID, "password").send_keys(password)

    # Locate and click login button element
    driver.find_element(By.ID, "login-button").click()
    time.sleep(3)

    # Locate and click menu button element
    driver.find_element(By.ID,"react-burger-menu-btn").click()
    time.sleep(3)

    # Locate and click logout button element
    driver.find_element(By.ID,"logout_sidebar_link").click()
    time.sleep(3)
    
driver.quit()